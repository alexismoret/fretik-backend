"""Reusable styling helpers for the xlsx skill.

Import via:
    import sys; sys.path.insert(0, "skills/xlsx/scripts")
    import workbook_helpers as wh
"""

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def style_header_row(
    ws: Worksheet,
    fill: str = "1F4E78",
    color: str = "FFFFFF",
    row: int = 1,
) -> None:
    """Apply the canonical header style to a row.

    Bold white text on deep-blue fill, centered vertically and
    horizontally. Matches the palette documented in SKILL.md.
    """
    header_font = Font(bold=True, color=color)
    header_fill = PatternFill("solid", fgColor=fill)
    header_align = Alignment(horizontal="center", vertical="center")
    for cell in ws[row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align


def autosize_columns(
    ws: Worksheet,
    min_w: int = 10,
    max_w: int = 60,
) -> None:
    """Set each column's width to the longest string length in it.

    Walks every cell once, tracks the max `len(str(value))` per column,
    clamps between `min_w` and `max_w`, and writes the width. Skips
    cells whose value is None.
    """
    widths: dict[int, int] = {}
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            col = cell.column
            if length > widths.get(col, 0):
                widths[col] = length
    for col, length in widths.items():
        width = max(min_w, min(length + 2, max_w))
        ws.column_dimensions[get_column_letter(col)].width = width


def add_source_footer(
    ws: Worksheet,
    lines: list[str],
    start_row: int | None = None,
) -> None:
    """Append italic-gray source lines one row below the current data.

    Pass the list of source strings in the format documented in
    SKILL.md: `Source: [System/Document], [Date], [Reference], [URL]`.

    If `start_row` is omitted, the footer starts two rows below
    `ws.max_row` (leaving a visual gap). Each line is merged across
    the used columns for cleanliness.
    """
    if not lines:
        return
    start = (start_row if start_row is not None else ws.max_row + 2)
    last_col = ws.max_column or 1
    footnote_font = Font(italic=True, color="808080", size=9)
    for i, line in enumerate(lines):
        row = start + i
        cell = ws.cell(row=row, column=1, value=line)
        cell.font = footnote_font
        if last_col > 1:
            ws.merge_cells(
                start_row=row,
                end_row=row,
                start_column=1,
                end_column=last_col,
            )
